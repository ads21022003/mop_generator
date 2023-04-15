#!/usr/bin/python
from openpyxl import load_workbook
from openpyxl import Workbook
from copy import copy
import smtplib,ssl
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.utils import formatdate
from email import encoders
from io import BytesIO
from io import StringIO
import nettools_api, getpass
from requests_ntlm import HttpNtlmAuth

def initial(path_variable):
    #path_variable=r"C:/Users/rajnkum.FAREAST/Desktop/MOPs/PLNET/DEMO/plnet_input_file.xlsx"
    variable_workbook = load_workbook(path_variable,data_only=True)
    sheet_variable = variable_workbook["Variable"]
    values_dict = {}
    for value in sheet_variable.iter_cols(min_row=2,min_col=4,max_col=4,values_only=True):
        temp = value
    for i in temp:
        temp_list = i.split(",")
        values_dict[temp_list[0]] = temp_list[1]
    #print(temp)
    #print(values_dict)
    return values_dict 

def sheetwriter(sheet,out_sheet,res):
    for x in range(1,4):
        for y in range(1,500):
            val = (sheet.cell(row=y,column=x)).value
            if val==None :
                out_sheet.cell(row=y,column=x).value = None
            else :
                val1 =val
                for p in res.keys():
                    if val1.find(p.strip()) !=-1:
                        val1=val1.replace(p.strip(),res[p])
                    out_sheet.cell(row=y,column=x).value = val1
                    
            if sheet.cell(row=y,column=x).has_style:
                out_sheet.cell(row=y,column=x).font = copy( sheet.cell(row=y,column=x).font)
                out_sheet.cell(row=y,column=x).border = copy( sheet.cell(row=y,column=x).border)
                out_sheet.cell(row=y,column=x).fill = copy( sheet.cell(row=y,column=x).fill)
                out_sheet.cell(row=y,column=x).number_format = copy( sheet.cell(row=y,column=x).number_format)
                out_sheet.cell(row=y,column=x).protection = copy( sheet.cell(row=y,column=x).protection)
                out_sheet.cell(row=y,column=x).alignment = copy( sheet.cell(row=y,column=x).alignment)


def send_mail(email,path):
    msg = MIMEMultipart()
    x = "PLNET MOP"+" for IAN and CORP peering on "+values_dict["<CORP_DEVICE_NAME_1>"]+" and "+values_dict["<CORP_DEVICE_NAME_1>"]
    mop_filename = "PLNET_MOP_FOR_IAN_CORP_"+values_dict["<CORP_DEVICE_NAME_1>"]+"_"+values_dict["<CORP_DEVICE_NAME_1>"]
    msg['From'] = "nistools@microsoft.com"
    msg['To'] = email
    msg['Subject'] = x
    part = MIMEBase('application', "octet-stream")
    part.set_payload(path.getvalue())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', 'attachment; filename = "'+mop_filename+'.xlsx"')
    msg.attach(part)


    server = smtplib.SMTP('10.169.23.229', 25)
    server.sendmail(msg['From'], msg['To'], msg.as_string())
    server.quit()
    print ("successfully sent email to %s:" % (msg['To']))
              

#template file initialization
while True:
    a =int(input("enter 1 for 1st tempelate and 2 for 2nd tempelate"))
    if a == 1:
        print('you have chosen 1st tempelate')
        path=r"C:\Users\rajnkum.FAREAST\Desktop\MOPs\PLNET\DEMO\plnet_mop_template.xlsx"
        path_variable=r"C:/Users/rajnkum.FAREAST/Desktop/MOPs/PLNET/DEMO/plnet_input_file.xlsx"
        break
    elif a == 2:
        print('you have chosen 2nd tempelate')
        path=r"C:\Users\rajnkum.FAREAST\Desktop\MOPs\PLNET\DEMO\plnet_mop_template_2.xlsx"
        path_variable=r"C:/Users/rajnkum.FAREAST/Desktop/MOPs/PLNET/DEMO/plnet_input_file_2.xlsx"
        break
   
    else:
        print('invalid input')
        
template = load_workbook(path)
sheet_project = template["Project"]
sheet_prepost = template["Pre_Post_Checks"]
sheet_corp = template["CORP"]
sheet_ian = template["IAN"]

#output file intialization
filename="output.xlsx"
path = BytesIO()
output = Workbook()
project_sheet = output.create_sheet("Project",0)
prepost_sheet = output.create_sheet("Pre_Post_Checks",1)
CORP_sheet = output.create_sheet("CORP",2)
IAN_sheet = output.create_sheet("IAN",3)
#calling input values from input file
values_dict = initial(path_variable)
#replacing value to create mop
sheetwriter(sheet_project,project_sheet,values_dict)
sheetwriter(sheet_prepost,prepost_sheet,values_dict)
sheetwriter(sheet_corp,CORP_sheet,values_dict)
sheetwriter(sheet_ian,IAN_sheet,values_dict)

#saving the file
# output.save(filename=path)

#mailing
email_id = input("Enter Your email as alias@microsoft.com in lower case: ")
#send_mail(email_id,path)

def pre_postcheck():
    temp = []
    for value in prepost_sheet.iter_cols(min_col=1,max_col=2,max_row=65,values_only=True):
        #print(value)
        command_list=[]
        for i in value[15:40]:
            if i !=None :
                if i[0] !='!':
                    command_list.append(i)
        
        temp.append(command_list)
        
    return temp
postcheck_choice = int(input("Do you want to perform Post Checks? If yes type numeric 1 else press any key. "))
        
if postcheck_choice == 1:
    temp = pre_postcheck()
    text = open("postcheck.txt", "w")

    #python function to call get_command_output returning is-is routing tables:
    devices_list = [values_dict['<CORP_DEVICE_NAME_1>'],values_dict['<CORP_DEVICE_NAME_1>']]
    user = input('Enter username: ')
    pw = getpass.getpass('Password: ')
    for i in range(2):
        output_list = list()

        device = devices_list[0]
        commands = temp[0]
        for command in commands:
            try:
                a =nettools_api.get_command_output(device, command, user, pw)
                text.write(command +'\n')
                text.write(a['output'])
                #for line in a['output'].split('\n'):
                #   print(line)

            except Exception as e:
                #raise Exception('netool_api module generated an exception: {}'.format(e))
                pass
else:
    print("Thankyou!!!")

send_mail(email_id,"postcheck.txt")
